from __future__ import annotations

import re
from io import BytesIO

import pandas as pd
import plotly.express as px
import streamlit as st

from st_aggrid import AgGrid, GridUpdateMode

from app_data import load_data, apply_filters
from data_utils import ringgit


APP_TITLE = "Outlet & Supplier Analysis"


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🏪",
    layout="wide",
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def create_empty_message() -> None:
    """
    Show a friendly message when filters return no data.
    """

    st.warning(
        "No rows match the current filters. "
        "Please select more values on the Dashboard page."
    )

    st.stop()


def sort_months(months: list) -> list:
    """
    Sort month values chronologically.

    Supports formats such as:

        MAR'26
        APR'26
        MAY'26

    and normal datetime-like values.
    """

    months = list(months)

    if not months:
        return []

    parsed = pd.to_datetime(
        months,
        format="%b'%y",
        errors="coerce",
    )

    if parsed.notna().all():

        return [
            month
            for _, month in sorted(
                zip(parsed, months),
                key=lambda item: item[0],
            )
        ]

    parsed = pd.to_datetime(
        months,
        errors="coerce",
    )

    if parsed.notna().all():

        return [
            month
            for _, month in sorted(
                zip(parsed, months),
                key=lambda item: item[0],
            )
        ]

    return sorted(
        months,
        key=str,
    )


def sku_search_text(sku: str) -> list[str]:
    """
    Convert a SKU/product string into searchable tokens.

    Example:

        207001001234-1-B 274 PRODUCTNAME 10KG

    becomes tokens such as:

        207001001234
        1
        B
        274
        PRODUCTNAME
        10KG
    """

    sku = str(sku).lower()

    tokens = re.split(
        r"[\s\-_\/]+",
        sku,
    )

    return [
        token.strip()
        for token in tokens
        if token.strip()
    ]


def search_skus(
    sku_values: list[str],
    search_text: str,
) -> list[str]:
    """
    Search SKU values using one or more search terms.

    A SKU is returned when ANY search term exists anywhere
    inside the full SKU text.
    """

    if not search_text.strip():
        return sku_values

    search_terms = re.split(
        r"[\s\-_\/]+",
        search_text.lower().strip(),
    )

    search_terms = [
        term
        for term in search_terms
        if term
    ]

    matched_skus = []

    for sku in sku_values:

        full_text = str(sku).lower()

        if any(
            term in full_text
            for term in search_terms
        ):

            matched_skus.append(sku)

    return matched_skus


def get_product_column(
    dataframe: pd.DataFrame,
) -> str | None:
    """
    Find the first available product-name column.
    """

    product_columns = [
        "PRODUCT",
        "PRODUCT_NAME",
        "DESCRIPTION",
        "ITEM",
        "ITEM_NAME",
        "SKU_NAME",
    ]

    return next(
        (
            column
            for column in product_columns
            if column in dataframe.columns
        ),
        None,
    )


def get_product_name(
    dataframe: pd.DataFrame,
) -> str:
    """
    Extract product name from the SKU column.

    Example:

        115003742-20 0023 LACTOGEN SUSU TEPUNG LANGKAH 1 20*350G

    becomes:

        LACTOGEN SUSU TEPUNG LANGKAH 1 20*350G
    """

    if "SKU" not in dataframe.columns:
        return "N/A"

    values = (
        dataframe["SKU"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    if not values:
        return "N/A"

    product_names = []

    for sku in values:

        match = re.match(
            r"^\S+\s+\d{4}\s+(.+)$",
            sku,
        )

        if match:
            product_name = match.group(1).strip()
        else:
            product_name = sku

        if product_name:
            product_names.append(product_name)

    if not product_names:
        return "N/A"

    return ", ".join(
        dict.fromkeys(product_names)
    )



def get_supplier_name(
    dataframe: pd.DataFrame,
) -> str:
    """
    Return a comma-separated list of suppliers.
    """

    if "SUPPLIER" not in dataframe.columns:
        return "N/A"

    values = (
        dataframe["SUPPLIER"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not values:
        return "N/A"

    return ", ".join(
        sorted(values)
    )


def build_outlet_monthly_table(
    dataframe: pd.DataFrame,
    available_months: list,
) -> pd.DataFrame:
    """
    Build the detailed SKU outlet monthly summary.

    Structure:

        NO
        WHOUSE
        OUTLET
        QTY
            month columns
            TOTAL
        AMOUNT
            month columns
            TOTAL
    """

    outlet_group_columns = []

    if "WHOUSE" in dataframe.columns:
        outlet_group_columns.append("WHOUSE")

    if "OUTLET" in dataframe.columns:
        outlet_group_columns.append("OUTLET")

    if not outlet_group_columns:
        raise ValueError(
            "The dataset needs at least "
            "'WHOUSE' or 'OUTLET' columns."
        )

# def convert_wide_monthly_data(
#     dataframe: pd.DataFrame,
# ) -> pd.DataFrame:
#     """
#     Convert wide monthly columns such as:

#         ~2026-02 QTY
#         ~2026-02 AMT
#         ~2026-03 QTY
#         ~2026-03 AMT

#     into normalized columns:

#         MONTH
#         QTY
#         AMT

#     TOTAL_QTY and TOTAL_AMT are NOT used as monthly data.
#     """

#     dataframe = dataframe.copy()

#     qty_pattern = re.compile(
#         r"^~?(\d{4})-(\d{2})\s+QTY$",
#         re.IGNORECASE,
#     )

#     amt_pattern = re.compile(
#         r"^~?(\d{4})-(\d{2})\s+AMT$",
#         re.IGNORECASE,
#     )

#     qty_month_columns = {}
#     amt_month_columns = {}

#     for column in dataframe.columns:

#         column_text = str(column).strip()

#         qty_match = qty_pattern.match(
#             column_text
#         )

#         if qty_match:

#             year = qty_match.group(1)
#             month = qty_match.group(2)

#             month_key = f"{year}-{month}"

#             qty_month_columns[
#                 month_key
#             ] = column

#         amt_match = amt_pattern.match(
#             column_text
#         )

#         if amt_match:

#             year = amt_match.group(1)
#             month = amt_match.group(2)

#             month_key = f"{year}-{month}"

#             amt_month_columns[
#                 month_key
#             ] = column

#     common_months = sorted(
#         set(qty_month_columns)
#         & set(amt_month_columns)
#     )

#     if not common_months:

#         raise ValueError(
#             "No monthly QTY/AMT columns found. "
#             "Expected columns such as '~2026-02 QTY' "
#             "and '~2026-02 AMT'."
#         )

#     id_columns = [
#         column
#         for column in dataframe.columns
#         if column not in (
#             list(qty_month_columns.values())
#             + list(amt_month_columns.values())
#             + [
#                 "TOTAL_QTY",
#                 "TOTAL_AMT",
#             ]
#         )
#     ]

#     output_rows = []

#     for month_key in common_months:

#         qty_column = qty_month_columns[
#             month_key
#         ]

#         amt_column = amt_month_columns[
#             month_key
#         ]

#         temp = dataframe[
#             id_columns
#             + [
#                 qty_column,
#                 amt_column,
#             ]
#         ].copy()

#         temp["MONTH"] = pd.to_datetime(
#             month_key + "-01",
#             format="%Y-%m-%d",
#         ).dt.strftime(
#             "%b'%y"
#         ).str.upper()

#         temp["QTY"] = pd.to_numeric(
#             temp[qty_column],
#             errors="coerce",
#         ).fillna(0)

#         temp["AMT"] = pd.to_numeric(
#             temp[amt_column],
#             errors="coerce",
#         ).fillna(0)

#         temp = temp.drop(
#             columns=[
#                 qty_column,
#                 amt_column,
#             ]
#         )

#         output_rows.append(temp)

#     normalized = pd.concat(
#         output_rows,
#         ignore_index=True,
#     )

#     return normalized

    # --------------------------------------------------------
    # QTY PIVOT
    # --------------------------------------------------------

    qty_pivot = (
        dataframe
        .pivot_table(
            index=outlet_group_columns,
            columns="MONTH",
            values="QTY",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    # --------------------------------------------------------
    # AMOUNT PIVOT
    # --------------------------------------------------------

    amount_pivot = (
        dataframe
        .pivot_table(
            index=outlet_group_columns,
            columns="MONTH",
            values="AMT",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    # --------------------------------------------------------
    # ENSURE ALL MONTHS EXIST
    # --------------------------------------------------------

    for month in available_months:

        if month not in qty_pivot.columns:
            qty_pivot[month] = 0

        if month not in amount_pivot.columns:
            amount_pivot[month] = 0

    # --------------------------------------------------------
    # KEEP CHRONOLOGICAL MONTH ORDER
    # --------------------------------------------------------

    qty_pivot = qty_pivot[
        outlet_group_columns
        + available_months
    ]

    amount_pivot = amount_pivot[
        outlet_group_columns
        + available_months
    ]

    # --------------------------------------------------------
    # RENAME QTY
    # --------------------------------------------------------

    qty_pivot = qty_pivot.rename(
        columns={
            month: f"QTY_{month}"
            for month in available_months
        }
    )

    # --------------------------------------------------------
    # RENAME AMOUNT
    # --------------------------------------------------------

    amount_pivot = amount_pivot.rename(
        columns={
            month: f"AMOUNT_{month}"
            for month in available_months
        }
    )

    # --------------------------------------------------------
    # MERGE
    # --------------------------------------------------------

    outlet_table = qty_pivot.merge(
        amount_pivot,
        on=outlet_group_columns,
        how="outer",
    )

    # --------------------------------------------------------
    # COLUMN LISTS
    # --------------------------------------------------------

    qty_columns = [
        f"QTY_{month}"
        for month in available_months
    ]

    amount_columns = [
        f"AMOUNT_{month}"
        for month in available_months
    ]

    # --------------------------------------------------------
    # TOTAL QTY
    # --------------------------------------------------------

    outlet_table["QTY_TOTAL"] = (
        outlet_table[qty_columns]
        .sum(axis=1)
    )

    # --------------------------------------------------------
    # TOTAL AMOUNT
    # --------------------------------------------------------

    outlet_table["AMOUNT_TOTAL"] = (
        outlet_table[amount_columns]
        .sum(axis=1)
    )

    # --------------------------------------------------------
    # SORT
    # --------------------------------------------------------

    sort_columns = [
        column
        for column in [
            "WHOUSE",
            "OUTLET",
        ]
        if column in outlet_table.columns
    ]

    if sort_columns:

        outlet_table = (
            outlet_table
            .sort_values(sort_columns)
            .reset_index(drop=True)
        )

    # --------------------------------------------------------
    # ADD NO
    # --------------------------------------------------------

    outlet_table.insert(
        0,
        "NO",
        range(
            1,
            len(outlet_table) + 1,
        ),
    )

    # --------------------------------------------------------
    # FINAL COLUMN ORDER
    # --------------------------------------------------------

    ordered_columns = ["NO"]

    if "WHOUSE" in outlet_table.columns:
        ordered_columns.append("WHOUSE")

    if "OUTLET" in outlet_table.columns:
        ordered_columns.append("OUTLET")

    ordered_columns += [
        f"QTY_{month}"
        for month in available_months
    ]

    ordered_columns.append(
        "QTY_TOTAL"
    )

    ordered_columns += [
        f"AMOUNT_{month}"
        for month in available_months
    ]

    ordered_columns.append(
        "AMOUNT_TOTAL"
    )

    return outlet_table[
        ordered_columns
    ]


def build_aggrid_options(
    dataframe: pd.DataFrame,
    available_months: list,
) -> dict:
    """
    Build AgGrid configuration for the detailed
    outlet monthly summary.
    """

    column_defs = []

    # --------------------------------------------------------
    # NO
    # --------------------------------------------------------

    column_defs.append(
        {
            "field": "NO",
            "headerName": "NO",
            "width": 70,
            "pinned": "left",
            "filter": "agNumberColumnFilter",
            "sortable": True,
        }
    )

    # --------------------------------------------------------
    # WHOUSE
    # --------------------------------------------------------

    if "WHOUSE" in dataframe.columns:

        column_defs.append(
            {
                "field": "WHOUSE",
                "headerName": "WHOUSE",
                "width": 100,
                "pinned": "left",
                "filter": "agTextColumnFilter",
                "sortable": True,
            }
        )

    # --------------------------------------------------------
    # OUTLET
    # --------------------------------------------------------

    if "OUTLET" in dataframe.columns:

        column_defs.append(
            {
                "field": "OUTLET",
                "headerName": "OUTLET",
                "width": 240,
                "pinned": "left",
                "filter": "agTextColumnFilter",
                "sortable": True,
            }
        )

    # ========================================================
    # QTY GROUP
    # ========================================================

    qty_children = []

    for month in available_months:

        qty_children.append(
            {
                "field": f"QTY_{month}",
                "headerName": str(month),
                "type": "numericColumn",
                "filter": "agNumberColumnFilter",
                "sortable": True,
                "width": 105,
                "valueFormatter": (
                    "params.value == null ? '' : "
                    "params.value.toLocaleString()"
                ),
            }
        )

    qty_children.append(
        {
            "field": "QTY_TOTAL",
            "headerName": "TOTAL",
            "type": "numericColumn",
            "filter": "agNumberColumnFilter",
            "sortable": True,
            "width": 115,
            "valueFormatter": (
                "params.value == null ? '' : "
                "params.value.toLocaleString()"
            ),
        }
    )

    column_defs.append(
        {
            "headerName": "QTY",
            "marryChildren": True,
            "children": qty_children,
        }
    )

    # ========================================================
    # AMOUNT GROUP
    # ========================================================

    amount_children = []

    for month in available_months:

        amount_children.append(
            {
                "field": f"AMOUNT_{month}",
                "headerName": str(month),
                "type": "numericColumn",
                "filter": "agNumberColumnFilter",
                "sortable": True,
                "width": 135,
                "valueFormatter": (
                    "params.value == null ? '' : "
                    "'RM ' + params.value.toLocaleString("
                    "undefined, "
                    "{minimumFractionDigits: 2, "
                    "maximumFractionDigits: 2}"
                    ")"
                ),
            }
        )

    amount_children.append(
        {
            "field": "AMOUNT_TOTAL",
            "headerName": "TOTAL",
            "type": "numericColumn",
            "filter": "agNumberColumnFilter",
            "sortable": True,
            "width": 145,
            "valueFormatter": (
                "params.value == null ? '' : "
                "'RM ' + params.value.toLocaleString("
                "undefined, "
                "{minimumFractionDigits: 2, "
                "maximumFractionDigits: 2}"
                ")"
            ),
        }
    )

    column_defs.append(
        {
            "headerName": "AMOUNT",
            "marryChildren": True,
            "children": amount_children,
        }
    )

    # ========================================================
    # GRID OPTIONS
    # ========================================================

    return {
        "columnDefs": column_defs,
        "defaultColDef": {
            "sortable": True,
            "filter": True,
            "resizable": True,
            "floatingFilter": True,
        },
        "pagination": True,
        "paginationPageSize": 20,
        "animateRows": True,
        "suppressHorizontalScroll": False,
        "sideBar": {
            "toolPanels": [
                "filters",
                "columns",
            ],
            "defaultToolPanel": "filters",
        },
        "enableRangeSelection": True,
        "statusBar": {
            "statusPanels": [
                {
                    "statusPanel": (
                        "agTotalAndFilteredRowCountComponent"
                    ),
                    "align": "left",
                },
                {
                    "statusPanel": (
                        "agAggregationComponent"
                    ),
                    "statusPanelParams": {
                        "aggFuncs": [
                            "sum",
                            "min",
                            "max",
                            "avg",
                        ]
                    },
                },
            ]
        },
    }


def create_excel_file(
    dataframe: pd.DataFrame,
    sku: str,
    supplier: str,
) -> bytes:
    """
    Create a formatted Excel workbook for the
    SKU outlet monthly summary.
    """

    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    output = BytesIO()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "SKU Outlet Summary"

    columns = list(
        dataframe.columns
    )

    # --------------------------------------------------------
    # Identify columns
    # --------------------------------------------------------

    qty_columns_excel = [
        index + 1
        for index, column in enumerate(columns)
        if column.startswith("QTY_")
    ]

    amount_columns_excel = [
        index + 1
        for index, column in enumerate(columns)
        if column.startswith("AMOUNT_")
    ]

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    worksheet.cell(
        row=1,
        column=1,
        value=(
            "SKU Outlet Monthly Summary\n"
            f"Supplier: {supplier}\n"
            f"SKU: {sku}"
        ),
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
    )

    title_cell.font = Font(
        bold=True,
        size=14,
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(columns),
    )

    worksheet.row_dimensions[1].height = 55

    # --------------------------------------------------------
    # HEADER ROWS
    # --------------------------------------------------------

    header_row_1 = 3
    header_row_2 = 4

    # --------------------------------------------------------
    # FIXED COLUMNS
    # --------------------------------------------------------

    fixed_columns = [
        column
        for column in columns
        if not column.startswith("QTY_")
        and not column.startswith("AMOUNT_")
    ]

    for column in fixed_columns:

        column_index = (
            columns.index(column) + 1
        )

        worksheet.cell(
            row=header_row_1,
            column=column_index,
            value=column,
        )

        worksheet.merge_cells(
            start_row=header_row_1,
            start_column=column_index,
            end_row=header_row_2,
            end_column=column_index,
        )

    # --------------------------------------------------------
    # QTY HEADER
    # --------------------------------------------------------

    if qty_columns_excel:

        qty_start = min(
            qty_columns_excel
        )

        qty_end = max(
            qty_columns_excel
        )

        worksheet.cell(
            row=header_row_1,
            column=qty_start,
            value="QTY",
        )

        worksheet.merge_cells(
            start_row=header_row_1,
            start_column=qty_start,
            end_row=header_row_1,
            end_column=qty_end,
        )

        for column_index in qty_columns_excel:

            column_name = columns[
                column_index - 1
            ]

            month_name = (
                column_name
                .replace("QTY_", "")
            )

            worksheet.cell(
                row=header_row_2,
                column=column_index,
                value=month_name,
            )

    # --------------------------------------------------------
    # AMOUNT HEADER
    # --------------------------------------------------------

    if amount_columns_excel:

        amount_start = min(
            amount_columns_excel
        )

        amount_end = max(
            amount_columns_excel
        )

        worksheet.cell(
            row=header_row_1,
            column=amount_start,
            value="AMOUNT",
        )

        worksheet.merge_cells(
            start_row=header_row_1,
            start_column=amount_start,
            end_row=header_row_1,
            end_column=amount_end,
        )

        for column_index in amount_columns_excel:

            column_name = columns[
                column_index - 1
            ]

            month_name = (
                column_name
                .replace("AMOUNT_", "")
            )

            worksheet.cell(
                row=header_row_2,
                column=column_index,
                value=month_name,
            )

    # --------------------------------------------------------
    # HEADER STYLING
    # --------------------------------------------------------

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    subheader_fill = PatternFill(
        "solid",
        fgColor="5B9BD5",
    )

    thin_border = Border(
        left=Side(
            style="thin",
            color="D9E1F2",
        ),
        right=Side(
            style="thin",
            color="D9E1F2",
        ),
        top=Side(
            style="thin",
            color="D9E1F2",
        ),
        bottom=Side(
            style="thin",
            color="D9E1F2",
        ),
    )

    for row in [
        header_row_1,
        header_row_2,
    ]:

        for column_index in range(
            1,
            len(columns) + 1,
        ):

            cell = worksheet.cell(
                row=row,
                column=column_index,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.border = thin_border

            if row == header_row_1:

                cell.fill = header_fill

            else:

                cell.fill = subheader_fill

    # --------------------------------------------------------
    # WRITE DATA
    # --------------------------------------------------------

    data_start_row = 5

    for row_number, row in enumerate(
        dataframe.itertuples(
            index=False
        ),
        start=data_start_row,
    ):

        for column_number, value in enumerate(
            row,
            start=1,
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

            cell.border = thin_border

            if column_number in qty_columns_excel:

                cell.number_format = "#,##0"

            elif column_number in amount_columns_excel:

                cell.number_format = (
                    '"RM "#,##0.00'
                )

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    for column_number, column_name in enumerate(
        columns,
        start=1,
    ):

        if column_name == "NO":

            width = 8

        elif column_name == "WHOUSE":

            width = 12

        elif column_name == "OUTLET":

            width = 28

        elif column_name.startswith("QTY_"):

            width = 14

        elif column_name.startswith("AMOUNT_"):

            width = 16

        else:

            width = 15

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = width

    # --------------------------------------------------------
    # FREEZE PANES
    # --------------------------------------------------------

    worksheet.freeze_panes = "D5"

    # --------------------------------------------------------
    # FILTER
    # --------------------------------------------------------

    worksheet.auto_filter.ref = (
        f"A4:"
        f"{get_column_letter(len(columns))}"
        f"{data_start_row + len(dataframe) - 1}"
    )

    # --------------------------------------------------------
    # ROW HEIGHTS
    # --------------------------------------------------------

    worksheet.row_dimensions[
        header_row_1
    ].height = 24

    worksheet.row_dimensions[
        header_row_2
    ].height = 22

    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    workbook.save(output)

    output.seek(0)

    return output.getvalue()


def build_outlet_summary(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:

    if "OUTLET" not in dataframe.columns:
        return pd.DataFrame(
            columns=[
                "OUTLET",
                "QTY",
                "AMT",
                "Sales Share %",
            ]
        )

    summary = (
        dataframe
        .groupby(
            "OUTLET",
            as_index=False,
        )
        .agg(
            QTY=("QTY", "sum"),
            AMT=("AMT", "sum"),
        )
    )

    total_sales = summary["AMT"].sum()

    if total_sales != 0:

        summary["Sales Share %"] = (
            summary["AMT"]
            / total_sales
            * 100
        )

    else:

        summary["Sales Share %"] = 0.0

    summary = (
        summary
        .sort_values(
            "AMT",
            ascending=False,
        )
        .reset_index(drop=True)
    )

    return summary



def build_supplier_summary(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build supplier sales summary.
    """

    if "SUPPLIER" not in dataframe.columns:

        return pd.DataFrame(
            columns=[
                "SUPPLIER",
                "QTY",
                "AMT",
                "Sales Share %",
            ]
        )

    summary = (
        dataframe
        .groupby(
            "SUPPLIER",
            as_index=False,
        )
        .agg(
            QTY=("QTY", "sum"),
            AMT=("AMT", "sum"),
        )
        .sort_values(
            "AMT",
            ascending=False,
        )
        .reset_index(drop=True)
    )

    total_sales = summary["AMT"].sum()

    if total_sales > 0:

        summary["Sales Share %"] = (
            summary["AMT"]
            .div(total_sales)
            .mul(100)
        )

    else:

        summary["Sales Share %"] = 0.0

    return summary


def style_summary_table(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Apply safe number formatting to summary tables.
    """

    format_dict = {}

    if "QTY" in dataframe.columns:
        format_dict["QTY"] = "{:,.0f}"

    if "AMT" in dataframe.columns:
        format_dict["AMT"] = "RM {:,.2f}"

    if "Sales Share %" in dataframe.columns:
        format_dict["Sales Share %"] = "{:.1f}%"

    return dataframe.style.format(
        format_dict
    )


# ============================================================
# SKU SELECTION
# ============================================================

def select_analysis_sku(
    filtered_data: pd.DataFrame,
) -> str:
    """
    Display searchable SKU selector and return
    the selected SKU.
    """

    st.subheader("🔎 Analysis Selection")

    if "SKU" not in filtered_data.columns:

        st.error(
            "The dataset does not contain a 'SKU' column."
        )

        st.stop()

    sku_values = sorted(
        filtered_data["SKU"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not sku_values:

        st.info(
            "No SKU/product data available."
        )

        st.stop()

    sku_search = st.text_input(
        "🔎 Search SKU / Product",
        placeholder=(
            "Type SKU, code, product name, etc."
        ),
        key="analysis_sku_search",
    )

    filtered_sku_options = search_skus(
        sku_values,
        sku_search,
    )

    if not filtered_sku_options:

        st.warning(
            f"No SKU found matching **{sku_search}**."
        )

        st.stop()

    selected_sku = st.selectbox(
        "Select ONE SKU / Product",
        options=filtered_sku_options,
        key="analysis_selected_sku",
    )

    return selected_sku


# ============================================================
# DETAILED SKU OUTLET MONTHLY SUMMARY
# ============================================================

def display_outlet_monthly_summary(
    analysis_data: pd.DataFrame,
    selected_sku: str,
) -> pd.DataFrame:
    """
    Display the detailed outlet/month summary.

    Returns the table so it can also be exported.
    """

    st.divider()

    st.subheader(
        "📋 SKU Outlet Monthly Summary"
    )

    st.caption(
        "Investigate the selected SKU's quantity and "
        "sales amount across warehouses and outlets."
    )

    supplier_name = get_supplier_name(
        analysis_data
    )

    product_name = get_product_name(
        analysis_data
    )

    st.info(
        f"**SKU:** {selected_sku}  |  "
        f"**Product:** {product_name}  |  "
        f"**Supplier:** {supplier_name}"
    )

    available_months = sort_months(
        analysis_data["MONTH"]
        .dropna()
        .unique()
        .tolist()
    )

    if not available_months:

        st.info(
            "No monthly data available for this SKU."
        )

        return pd.DataFrame()

    try:

        outlet_table = build_outlet_monthly_table(
            analysis_data,
            available_months,
        )

    except ValueError as error:

        st.error(str(error))

        return pd.DataFrame()

    # --------------------------------------------------------
    # AgGrid
    # --------------------------------------------------------

    grid_options = build_aggrid_options(
        outlet_table,
        available_months,
    )

    AgGrid(
        outlet_table,
        gridOptions=grid_options,
        update_mode=GridUpdateMode.NO_UPDATE,
        theme="streamlit",
        height=600,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        key="single_sku_outlet_monthly_summary",
    )

    return outlet_table


# ============================================================
# EXCEL EXPORT
# ============================================================

def display_export_section(
    outlet_table: pd.DataFrame,
    selected_sku: str,
    supplier_name: str,
) -> None:
    """
    Display Excel download button.
    """

    if outlet_table.empty:
        return

    st.divider()

    st.subheader("📥 Export")

    st.caption(
        "Download the selected SKU outlet summary "
        "as a formatted Excel file."
    )

    excel_file = create_excel_file(
        outlet_table,
        selected_sku,
        supplier_name,
    )

    safe_sku = re.sub(
        r"[^A-Za-z0-9_\-]+",
        "_",
        str(selected_sku),
    )

    st.download_button(
        label="📥 Download Excel",
        data=excel_file,
        file_name=(
            f"SKU_{safe_sku}_"
            "Outlet_Monthly_Summary.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key="download_single_sku_outlet_excel",
    )


# ============================================================
# OUTLET PERFORMANCE
# ============================================================

def display_outlet_performance(
    analysis_data: pd.DataFrame,
    selected_sku: str,
) -> pd.DataFrame:
    """
    Display outlet KPI, ranking and monthly performance.
    """

    st.divider()

    st.subheader(
        f"🏪 Outlet Performance — {selected_sku}"
    )

    outlet_summary = build_outlet_summary(
        analysis_data
    )

    if outlet_summary.empty:

        st.info(
            "No outlet data is available for this SKU."
        )

        return outlet_summary

    # ========================================================
    # KPI
    # ========================================================

    total_amount = analysis_data["AMT"].sum()

    total_qty = analysis_data["QTY"].sum()

    active_outlets = analysis_data[
        "OUTLET"
    ].nunique()

    top_outlet = (
        outlet_summary.iloc[0]["OUTLET"]
    )

    st.markdown("**📌 Outlet KPIs**")

    kpi = st.columns(4)

    kpi[0].metric(
        "Total Sales",
        ringgit(total_amount),
    )

    kpi[1].metric(
        "Units Sold",
        f"{total_qty:,.0f}",
    )

    kpi[2].metric(
        "Active Outlets",
        f"{active_outlets:,}",
    )

    kpi[3].metric(
        "Top Outlet",
        str(top_outlet),
    )

    # ========================================================
    # RANKING
    # ========================================================

    st.markdown("**🏆 Outlet Sales Ranking**")

    chart_data = (
        outlet_summary
        .sort_values(
            "AMT",
            ascending=True,
        )
    )

    outlet_chart = px.bar(
        chart_data,
        x="AMT",
        y="OUTLET",
        orientation="h",
        labels={
            "AMT": "Sales Amount (RM)",
            "OUTLET": "Outlet",
        },
        color="AMT",
        color_continuous_scale="Teal",
    )

    outlet_chart.update_layout(
        coloraxis_showscale=False,
        height=max(
            400,
            min(
                700,
                len(chart_data) * 28,
            ),
        ),
        margin=dict(
            l=10,
            r=10,
            t=20,
            b=20,
        ),
        xaxis_title="Sales (RM)",
        yaxis_title="",
    )

    st.plotly_chart(
        outlet_chart,
        use_container_width=True,
        key="outlet_sales_ranking_chart",
    )

    # ========================================================
    # TABLE
    # ========================================================

    st.dataframe(
        style_summary_table(
            outlet_summary
        ),
        use_container_width=True,
        hide_index=True,
    )

    # ========================================================
    # MONTHLY PERFORMANCE
    # ========================================================

    display_outlet_monthly_performance(
        analysis_data,
        selected_sku,
    )

    return outlet_summary


def display_outlet_monthly_performance(
    analysis_data: pd.DataFrame,
    selected_sku: str,
) -> None:
    """
    Display overall and selected-outlet monthly sales.
    """

    st.subheader(
        "📈 Outlet Monthly Performance"
    )

    outlet_monthly = (
        analysis_data
        .groupby(
            ["MONTH", "OUTLET"],
            as_index=False,
        )["AMT"]
        .sum()
    )

    outlet_monthly_total = (
        analysis_data
        .groupby(
            "MONTH",
            as_index=False,
        )["AMT"]
        .sum()
    )

    available_months = sort_months(
        outlet_monthly_total[
            "MONTH"
        ].tolist()
    )

    outlet_monthly_total["MONTH_ORDER"] = (
        outlet_monthly_total["MONTH"]
        .map(
            {
                month: index
                for index, month
                in enumerate(available_months)
            }
        )
    )

    outlet_monthly_total = (
        outlet_monthly_total
        .sort_values("MONTH_ORDER")
        .drop(
            columns="MONTH_ORDER"
        )
    )

    if len(outlet_monthly_total) < 2:

        st.info(
            "At least two months are required "
            "for monthly outlet analysis."
        )

        return

    # ========================================================
    # OVERALL MONTHLY TREND
    # ========================================================

    monthly_chart = px.line(
        outlet_monthly_total,
        x="MONTH",
        y="AMT",
        markers=True,
        text="AMT",
        labels={
            "MONTH": "Month",
            "AMT": "Sales Amount (RM)",
        },
    )

    monthly_chart.update_traces(
        line=dict(
            color="#0F766E",
            width=3,
        ),
        marker=dict(
            size=9,
        ),
        texttemplate=(
            "RM %{text:,.0f}"
        ),
        textposition="top center",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Sales: RM %{y:,.2f}"
            "<extra></extra>"
        ),
    )

    monthly_chart.update_layout(
        height=400,
        hovermode="x unified",
        showlegend=False,
        margin=dict(
            l=20,
            r=20,
            t=20,
            b=20,
        ),
        xaxis_title="",
        yaxis_title="Sales (RM)",
    )

    st.plotly_chart(
        monthly_chart,
        use_container_width=True,
        key="outlet_overall_monthly_chart",
    )

    # ========================================================
    # OUTLET SELECTOR
    # ========================================================

    available_outlets = sorted(
        analysis_data["OUTLET"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not available_outlets:
        return

    selected_outlet = st.selectbox(
        "Select Outlet to investigate",
        available_outlets,
        key="outlet_analysis_selected_outlet",
    )

    selected_outlet_data = (
        outlet_monthly[
            outlet_monthly["OUTLET"].astype(str)
            == str(selected_outlet)
        ]
        .copy()
    )

    selected_months = sort_months(
        selected_outlet_data[
            "MONTH"
        ].tolist()
    )

    selected_outlet_data[
        "MONTH_ORDER"
    ] = selected_outlet_data[
        "MONTH"
    ].map(
        {
            month: index
            for index, month
            in enumerate(selected_months)
        }
    )

    selected_outlet_data = (
        selected_outlet_data
        .sort_values("MONTH_ORDER")
        .drop(
            columns="MONTH_ORDER"
        )
    )

    outlet_trend_chart = px.line(
        selected_outlet_data,
        x="MONTH",
        y="AMT",
        markers=True,
        text="AMT",
        labels={
            "MONTH": "Month",
            "AMT": "Sales Amount (RM)",
        },
        title=(
            f"Monthly Sales: "
            f"{selected_outlet}"
        ),
    )

    outlet_trend_chart.update_traces(
        line=dict(
            color="#0F766E",
            width=3,
        ),
        marker=dict(
            size=9,
        ),
        texttemplate=(
            "RM %{text:,.0f}"
        ),
        textposition="top center",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Sales: RM %{y:,.2f}"
            "<extra></extra>"
        ),
    )

    outlet_trend_chart.update_layout(
        height=400,
        hovermode="x unified",
        showlegend=False,
        margin=dict(
            l=20,
            r=20,
            t=50,
            b=20,
        ),
    )

    st.plotly_chart(
        outlet_trend_chart,
        use_container_width=True,
        key="outlet_selected_trend_chart",
    )


# ============================================================
# SUPPLIER PERFORMANCE
# ============================================================

def display_supplier_performance(
    analysis_data: pd.DataFrame,
    selected_sku: str,
) -> pd.DataFrame:
    """
    Display supplier KPI, ranking and monthly performance.
    """

    st.divider()

    st.subheader(
        f"📦 Supplier Performance — {selected_sku}"
    )

    supplier_summary = build_supplier_summary(
        analysis_data
    )

    if supplier_summary.empty:

        st.info(
            "No supplier data is available "
            "for this SKU."
        )

        return supplier_summary

    # ========================================================
    # KPI
    # ========================================================

    total_amount = analysis_data["AMT"].sum()

    total_qty = analysis_data["QTY"].sum()

    active_suppliers = analysis_data[
        "SUPPLIER"
    ].nunique()

    top_supplier = (
        supplier_summary.iloc[0]["SUPPLIER"]
    )

    st.markdown("**📌 Supplier KPIs**")

    kpi = st.columns(4)

    kpi[0].metric(
        "Total Sales",
        ringgit(total_amount),
    )

    kpi[1].metric(
        "Units Sold",
        f"{total_qty:,.0f}",
    )

    kpi[2].metric(
        "Active Suppliers",
        f"{active_suppliers:,}",
    )

    kpi[3].metric(
        "Top Supplier",
        str(top_supplier),
    )

    # ========================================================
    # RANKING
    # ========================================================

    st.markdown("**🏆 Supplier Sales Ranking**")

    chart_data = (
        supplier_summary
        .sort_values(
            "AMT",
            ascending=True,
        )
    )

    supplier_chart = px.bar(
        chart_data,
        x="AMT",
        y="SUPPLIER",
        orientation="h",
        labels={
            "AMT": "Sales Amount (RM)",
            "SUPPLIER": "Supplier",
        },
        color="AMT",
        color_continuous_scale="Blues",
    )

    supplier_chart.update_layout(
        coloraxis_showscale=False,
        height=max(
            400,
            min(
                700,
                len(chart_data) * 28,
            ),
        ),
        margin=dict(
            l=10,
            r=10,
            t=20,
            b=20,
        ),
        xaxis_title="Sales (RM)",
        yaxis_title="",
    )

    st.plotly_chart(
        supplier_chart,
        use_container_width=True,
        key="supplier_sales_ranking_chart",
    )

    # ========================================================
    # TABLE
    # ========================================================

    st.dataframe(
        style_summary_table(
            supplier_summary
        ),
        use_container_width=True,
        hide_index=True,
    )

    # ========================================================
    # MONTHLY PERFORMANCE
    # ========================================================

    display_supplier_monthly_performance(
        analysis_data,
        selected_sku,
    )

    return supplier_summary


def display_supplier_monthly_performance(
    analysis_data: pd.DataFrame,
    selected_sku: str,
) -> None:
    """
    Display overall and selected-supplier monthly sales.
    """

    st.subheader(
        "📈 Supplier Monthly Performance"
    )

    supplier_monthly = (
        analysis_data
        .groupby(
            ["MONTH", "SUPPLIER"],
            as_index=False,
        )["AMT"]
        .sum()
    )

    supplier_monthly_total = (
        analysis_data
        .groupby(
            "MONTH",
            as_index=False,
        )["AMT"]
        .sum()
    )

    available_months = sort_months(
        supplier_monthly_total[
            "MONTH"
        ].tolist()
    )

    supplier_monthly_total[
        "MONTH_ORDER"
    ] = (
        supplier_monthly_total["MONTH"]
        .map(
            {
                month: index
                for index, month
                in enumerate(available_months)
            }
        )
    )

    supplier_monthly_total = (
        supplier_monthly_total
        .sort_values("MONTH_ORDER")
        .drop(
            columns="MONTH_ORDER"
        )
    )

    if len(supplier_monthly_total) < 2:

        st.info(
            "At least two months are required "
            "for monthly supplier analysis."
        )

        return

    # ========================================================
    # OVERALL MONTHLY TREND
    # ========================================================

    supplier_total_chart = px.line(
        supplier_monthly_total,
        x="MONTH",
        y="AMT",
        markers=True,
        text="AMT",
        labels={
            "MONTH": "Month",
            "AMT": "Sales Amount (RM)",
        },
    )

    supplier_total_chart.update_traces(
        line=dict(
            color="#2563EB",
            width=3,
        ),
        marker=dict(
            size=9,
        ),
        texttemplate=(
            "RM %{text:,.0f}"
        ),
        textposition="top center",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Sales: RM %{y:,.2f}"
            "<extra></extra>"
        ),
    )

    supplier_total_chart.update_layout(
        height=400,
        hovermode="x unified",
        showlegend=False,
        margin=dict(
            l=20,
            r=20,
            t=20,
            b=20,
        ),
        xaxis_title="",
        yaxis_title="Sales (RM)",
    )

    st.plotly_chart(
        supplier_total_chart,
        use_container_width=True,
        key="supplier_overall_monthly_chart",
    )

    # ========================================================
    # SUPPLIER SELECTOR
    # ========================================================

    available_suppliers = sorted(
        analysis_data["SUPPLIER"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not available_suppliers:
        return

    selected_supplier = st.selectbox(
        "Select Supplier to investigate",
        available_suppliers,
        key="supplier_analysis_selected_supplier",
    )

    selected_supplier_data = (
        supplier_monthly[
            supplier_monthly["SUPPLIER"].astype(str)
            == str(selected_supplier)
        ]
        .copy()
    )

    selected_months = sort_months(
        selected_supplier_data[
            "MONTH"
        ].tolist()
    )

    selected_supplier_data[
        "MONTH_ORDER"
    ] = selected_supplier_data[
        "MONTH"
    ].map(
        {
            month: index
            for index, month
            in enumerate(selected_months)
        }
    )

    selected_supplier_data = (
        selected_supplier_data
        .sort_values("MONTH_ORDER")
        .drop(
            columns="MONTH_ORDER"
        )
    )

    supplier_trend_chart = px.line(
        selected_supplier_data,
        x="MONTH",
        y="AMT",
        markers=True,
        text="AMT",
        labels={
            "MONTH": "Month",
            "AMT": "Sales Amount (RM)",
        },
        title=(
            f"Monthly Sales: "
            f"{selected_supplier}"
        ),
    )

    supplier_trend_chart.update_traces(
        line=dict(
            color="#2563EB",
            width=3,
        ),
        marker=dict(
            size=9,
        ),
        texttemplate=(
            "RM %{text:,.0f}"
        ),
        textposition="top center",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Sales: RM %{y:,.2f}"
            "<extra></extra>"
        ),
    )

    supplier_trend_chart.update_layout(
        height=400,
        hovermode="x unified",
        showlegend=False,
        margin=dict(
            l=20,
            r=20,
            t=50,
            b=20,
        ),
    )

    st.plotly_chart(
        supplier_trend_chart,
        use_container_width=True,
        key="supplier_selected_trend_chart",
    )

    # ============================================================
    # MANAGEMENT INSIGHTS
    # ============================================================

def display_management_insights(
    outlet_summary: pd.DataFrame,
    supplier_summary: pd.DataFrame,
    selected_sku: str,
) -> None:
    """
    Display simple and significant management insights
    for the selected SKU.

    Logic:
    - Always show the top outlet/supplier.
    - If more than 1 outlet/supplier exists, also show
      the lowest performer.
    - Show Top 5 contribution when there is more than
      1 outlet/supplier.
    """

    st.divider()

    st.subheader(
        f"💡 Management Insights — {selected_sku}"
    )

    insight_columns = st.columns(2)

    # ========================================================
    # OUTLET INSIGHTS
    # ========================================================

    with insight_columns[0]:

        st.markdown("**🏪 Outlet Insights**")

        if outlet_summary.empty:

            st.info(
                "No outlet insights available."
            )

        else:

            best_outlet = outlet_summary.iloc[0]

            outlet_count = len(
                outlet_summary
            )

            # ------------------------------------------------
            # TOP OUTLET
            # ------------------------------------------------

            st.success(
                f"🏆 **Top outlet:** "
                f"{best_outlet['OUTLET']} generated "
                f"{ringgit(best_outlet['AMT'])}, "
                f"contributing "
                f"**{best_outlet['Sales Share %']:.1f}%** "
                f"of total sales."
            )

            # ------------------------------------------------
            # MULTIPLE OUTLETS
            # ------------------------------------------------

            if outlet_count > 1:

                worst_outlet = (
                    outlet_summary.iloc[-1]
                )

                st.warning(
                    f"⚠️ **Lowest-sales outlet:** "
                    f"{worst_outlet['OUTLET']} generated "
                    f"{ringgit(worst_outlet['AMT'])}."
                )

                # --------------------------------------------
                # TOP 5 CONTRIBUTION
                # --------------------------------------------

                top_5_outlet_share = (
                    outlet_summary
                    .head(5)["Sales Share %"]
                    .sum()
                )

                st.info(
                    f"📊 **Top {min(5, outlet_count)} "
                    f"outlets** contribute "
                    f"**{top_5_outlet_share:.1f}%** "
                    f"of total sales."
                )


    # ========================================================
    # SUPPLIER INSIGHTS
    # ========================================================

    with insight_columns[1]:

        st.markdown("**📦 Supplier Insights**")

        if supplier_summary.empty:

            st.info(
                "No supplier insights available."
            )

        else:

            best_supplier = supplier_summary.iloc[0]

            supplier_count = len(
                supplier_summary
            )

            # ------------------------------------------------
            # TOP SUPPLIER
            # ------------------------------------------------

            st.success(
                f"🏆 **Top supplier:** "
                f"{best_supplier['SUPPLIER']} generated "
                f"{ringgit(best_supplier['AMT'])}, "
                f"contributing "
                f"**{best_supplier['Sales Share %']:.1f}%** "
                f"of total sales."
            )

            # ------------------------------------------------
            # MULTIPLE SUPPLIERS
            # ------------------------------------------------

            if supplier_count > 1:

                worst_supplier = (
                    supplier_summary.iloc[-1]
                )

                st.warning(
                    f"⚠️ **Lowest-sales supplier:** "
                    f"{worst_supplier['SUPPLIER']} generated "
                    f"{ringgit(worst_supplier['AMT'])}."
                )

                # --------------------------------------------
                # TOP 5 CONTRIBUTION
                # --------------------------------------------

                top_5_supplier_share = (
                    supplier_summary
                    .head(5)["Sales Share %"]
                    .sum()
                )

                st.info(
                    f"📊 **Top {min(5, supplier_count)} "
                    f"suppliers** contribute "
                    f"**{top_5_supplier_share:.1f}%** "
                    f"of total sales."
                )



# ============================================================
# MAIN PAGE
# ============================================================

def main() -> None:

    # ========================================================
    # HEADER
    # ========================================================

    st.title(
        "🏪📦 Outlet & Supplier Analysis"
    )

    st.caption(
        "Understand which outlets and suppliers are driving "
        "sales performance for the selected SKU."
    )

    # ========================================================
    # LOAD DATA
    # ========================================================

    data = load_data()

    filtered_data = apply_filters(
        data
    )

    if filtered_data.empty:
        create_empty_message()

    # # Convert wide monthly columns into:
    # # MONTH / QTY / AMT
    # filtered_data = convert_wide_monthly_data(
    #     filtered_data
    # )

    # ========================================================
    # FILTER CONTEXT
    # ========================================================

    st.info(
        f"Showing **{len(filtered_data):,}** rows "
        f"after the current dashboard filters."
    )

    # ========================================================
    # SKU SELECTION
    # ========================================================

    selected_sku = select_analysis_sku(
        filtered_data
    )

    # ========================================================
    # SELECTED SKU DATA
    # ========================================================

    analysis_data = filtered_data[
        filtered_data["SKU"].astype(str)
        == str(selected_sku)
    ].copy()

    if analysis_data.empty:

        st.warning(
            "No data found for the selected SKU."
        )

        st.stop()

    # ========================================================
    # PRODUCT CONTEXT
    # ========================================================

    product_name = get_product_name(
        analysis_data
    )

    supplier_name = get_supplier_name(
        analysis_data
    )

    st.success(
        f"🔎 **Selected product:** {product_name}  |  "
        f"**SKU:** {selected_sku}  |  "
        f"**Supplier:** {supplier_name}"
    )

    # ========================================================
    # DETAILED MONTHLY TABLE
    # ========================================================

    outlet_table = (
        display_outlet_monthly_summary(
            analysis_data,
            selected_sku,
        )
    )

    # ========================================================
    # EXPORT
    # ========================================================

    display_export_section(
        outlet_table,
        selected_sku,
        supplier_name,
    )

    # ========================================================
    # OUTLET PERFORMANCE
    # ========================================================

    outlet_summary = (
        display_outlet_performance(
            analysis_data,
            selected_sku,
        )
    )

    # ========================================================
    # SUPPLIER PERFORMANCE
    # ========================================================

    supplier_summary = (
        display_supplier_performance(
            analysis_data,
            selected_sku,
        )
    )

    # ========================================================
    # MANAGEMENT INSIGHTS
    # ========================================================

    display_management_insights(
        outlet_summary,
        supplier_summary,
        selected_sku,
    )


# ============================================================
# RUN APP
# ============================================================

if __name__ == "__main__":
    main()